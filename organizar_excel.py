import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = "resultado_detran_organizado.xlsx"
SHEET_NAME = "Resultado DETRAN"


def formatar_excel():
    # ================= VALIDAÇÃO =================
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Arquivo não encontrado: {EXCEL_FILE}")
        return

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
        print(f"✅ Excel lido com sucesso: {EXCEL_FILE}")
    except Exception as e:
        print(f"❌ Erro ao ler Excel: {e}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # ================= ESTILOS =================
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ================= CABEÇALHO =================
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # ================= DADOS =================
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

            # Colunas com texto longo
            if cell.column in (5, 11):  # Motivo | Código de Barras
                cell.alignment = left
            else:
                cell.alignment = center

    # ================= LARGURA COLUNAS =================
    larguras = {
        "A": 12,  # Placa
        "B": 5,   # #
        "C": 15,  # AIT
        "D": 18,  # AIT Originária
        "E": 55,  # Motivo
        "F": 14,  # Data Infração
        "G": 16,  # Data Vencimento
        "H": 16,  # Valor
        "I": 18,  # Valor a Pagar
        "J": 20,  # Órgão Autuador
        "K": 55,  # Código de barras
    }

    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    # ================= EXTRAS =================
    ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)

    print("\n✅ Excel formatado com sucesso!")
    print(f"📄 Arquivo: {EXCEL_FILE}")
    print(f"📊 Total de multas: {len(df)}")
    print(f"📋 Colunas: {', '.join(df.columns)}")


if __name__ == "__main__":
    formatar_excel()
